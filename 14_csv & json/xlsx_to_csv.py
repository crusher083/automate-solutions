import os
import csv
import openpyxl

for excel in os.listdir('D://Python'):
    # Skip non-xlsx files, load the workbook object.
    if not excel.endswith('.xlsx'):
        continue
    wb = openpyxl.load_workbook(excel)
    for sheet_name in wb.sheetnames:
        # Loop through every sheet in the workbook.
        sheet = wb[sheet_name]
        # Create the CSV filename from the Excel filename and sheet title.
        csv_file = open(f'{excel[:-5]}_{sheet_name}.csv', 'w', newline='')
        # Create the csv.writer object for this CSV file.
        csv_wrt = csv.writer(csv_file)
        # Loop through every row in the sheet.
        for row in range(1, sheet.max_row + 1):
            row_data = []  # append each cell to this list
            for col in range(1, sheet.max_column + 1):
                row_data.append(sheet.cell(row=row, column=col).value)
        # Write the rowData list to the CSV file.
            for r in row_data:
                csv_wrt.writerow([r])
        csv_file.close()
