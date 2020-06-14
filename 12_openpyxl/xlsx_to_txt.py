import openpyxl

wb = openpyxl.load_workbook('new_xlsx.xlsx')
sheet = wb.active
rows = sheet.max_row
cols = sheet.max_column
for n in range(1, cols + 1):
    with open (f'new_xlsx{n}.txt', 'w') as f:
        for m in range(1, rows + 1):
            f.write(str(sheet.cell(row=m, column=n).value) + '\n')
wb.close()