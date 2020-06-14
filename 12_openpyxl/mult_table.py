import openpyxl
from openpyxl.styles import Font
import sys

wb = openpyxl.Workbook()
sheet = wb.active
n = int(sys.argv[1])
for i in range(1, n + 1):
    bold = Font(bold=True)
    sheet.cell(row=i + 1, column=1).value = i
    sheet.cell(row=i + 1, column=1).font = bold
    sheet.cell(row=1, column=(i + 1)).value = i
    sheet.cell(row=1, column=(i + 1)).font = bold
    for col in range(1, n + 1):
        sheet.cell(row=i + 1, column=col + 1).value = i * col

wb.save(f'multiplication{n}.xlsx')