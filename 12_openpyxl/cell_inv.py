import openpyxl

wb = openpyxl.load_workbook('inverse.xlsx')
sheet = wb.active
rows = sheet.max_row
cols = sheet.max_column
invert = wb.create_sheet("Inverted")
for n in range(1, rows + 1):
    for m in range(1, cols + 1):
        invert.cell(row=m, column=n).value = sheet.cell(row=n, column=m).value
wb.save(f'inversed.xlsx')
